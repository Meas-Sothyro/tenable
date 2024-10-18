import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def save_to_excel(combined_va_df: pd.DataFrame, combined_hardening_df: pd.DataFrame, output_filename: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if not combined_va_df.empty:
        ws_va = wb.create_sheet(title="VA")
        style_va_worksheet(ws_va, combined_va_df)

    if not combined_hardening_df.empty:
        ws_hardening = wb.create_sheet(title="Hardening")
        style_hardening_worksheet(ws_hardening, combined_hardening_df)

    wb.save(output_filename)

def style_va_worksheet(worksheet, df):
    # Add headers
    worksheet.append(df.columns.tolist())

    # Apply header styling
    header_fill = PatternFill(start_color="0F243E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Append data rows
    for row in df.itertuples(index=False, name=None):
        worksheet.append(row)

    # Define severity styles
    severity_styles = {
        "critical": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        "high": PatternFill(start_color="FE0000", end_color="FE0000", fill_type="solid"),
        "medium": PatternFill(start_color="FFC001", end_color="FFC001", fill_type="solid"),
        "low": PatternFill(start_color="72AC48", end_color="72AC48", fill_type="solid"),
        "info": PatternFill(start_color="3074B5", end_color="3074B5", fill_type="solid"),
    }

    # Apply conditional formatting based on severity
    severity_column_index = df.columns.get_loc("Severity") + 1
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=severity_column_index, max_col=severity_column_index):
        for cell in row:
            severity = str(cell.value).strip().lower()
            if severity in severity_styles:
                cell.fill = severity_styles[severity]
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")

    # Adjust column widths with a maximum limit
    adjust_column_widths(worksheet, df)

def style_hardening_worksheet(worksheet, df):
    # Add headers
    worksheet.append(df.columns.tolist())

    # Apply header styling
    header_fill = PatternFill(start_color="0F243E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Append data rows
    for row in df.itertuples(index=False, name=None):
        worksheet.append(row)

    # Apply conditional formatting based on Severity for Hardening
    severity_column_index = df.columns.get_loc("Severity") + 1
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=severity_column_index, max_col=severity_column_index):
        for cell in row:
            if cell.value == 'FAILED':
                cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
            elif cell.value == 'Manual Review':
                cell.fill = PatternFill(start_color='FFC100', end_color='FFC100', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
            elif cell.value == 'PASSED':
                cell.fill = PatternFill(start_color='3074B5', end_color='3074B5', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')

    # Adjust column widths with a maximum limit
    adjust_column_widths(worksheet, df)

def adjust_column_widths(worksheet, df, max_width=40):
    """
    Adjusts the column widths based on the content, with a maximum width limit.
    """
    for col in df.columns:
        col_letter = get_column_letter(df.columns.get_loc(col) + 1)
        max_length = max((len(str(cell.value)) for cell in worksheet[col_letter]), default=10) + 2
        if max_length > max_width:
            max_length = max_width  # Limit the column width
        worksheet.column_dimensions[col_letter].width = max_length